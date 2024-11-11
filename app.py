import os
import pandas as pd
import streamlit as st
from io import BytesIO
from datetime import datetime
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from cryptography.fernet import Fernet

# Version and Update Information
SCRIPT_VERSION = "v1.2"
UPDATE_DATE = "2024-11-01"

# Generate a key for encryption
encryption_key = Fernet.generate_key()
cipher = Fernet(encryption_key)

# Custom CSS for UI enhancements
def set_custom_style():
    st.markdown(
        f"""
        <style>
            body {{
                background-color: #FAF3E0;
                color: #333;
            }}
            h1, h2, h3, h4, h5, h6 {{
                color: #003366;
            }}
            .stButton > button {{
                background-color: #003366;
                color: white;
                border-radius: 12px;
                padding: 10px 20px;
                font-size: 16px;
            }}
            .stButton > button:hover {{
                background-color: #00509E;
            }}
            .stFileUploader {{
                border: 1px solid #ccc;
                padding: 10px;
                border-radius: 8px;
                background-color: #FAF3E0;
            }}
        </style>
        <div style="font-size: small; text-align: right; color: #888;">
            <p>Script Version: {SCRIPT_VERSION}</p>
            <p>Last Update: {UPDATE_DATE}</p>
        </div>
        """,
        unsafe_allow_html=True
    )

def encrypt_file(file_data):
    """Encrypt the uploaded file."""
    return cipher.encrypt(file_data)

def decrypt_file(encrypted_data):
    """Decrypt the encrypted file in memory."""
    return cipher.decrypt(encrypted_data)

# Custom report generation with styling
def create_styled_report(df):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # Style variables
    header_font = Font(bold=True, size=13)
    regular_font = Font(size=12)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style="thin", color="333333"),
                         right=Side(style="thin", color="333333"),
                         top=Side(style="thin", color="333333"),
                         bottom=Side(style="thin", color="333333"))
    thick_border = Border(left=Side(style="thick", color="333333"),
                          right=Side(style="thick", color="333333"),
                          top=Side(style="thick", color="333333"),
                          bottom=Side(style="thick", color="333333"))

    # Set column headers and arrange them based on the specified order
    ordered_columns = ['تاریخ', 'کارت به کارت', 'فروش', 'مالیات', 'کارمزد', 'برداشت روز', 'مانده آخر روز', 'واریزی اسنپ']
    df = df[ordered_columns]

    # Populate table with headers and data, setting column headers first with styling
    ws.append(ordered_columns)
    for cell in ws[1]:  # Apply header styling
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thick_border
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    # Populate data rows with formatting
    for row in df.itertuples(index=False, name=None):
        ws.append(row)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = regular_font
            cell.alignment = center_align
            cell.border = thin_border
            # Apply number format
            if cell.column_letter != 'A':  # Skip date column
                cell.number_format = '#,##0.00'
            else:
                cell.number_format = '[$-fa-IR,700]yyyy/mm/dd;@'  # Shamsi date format for Persian locale

    # Set column widths and row heights
    for col in ws.columns:
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = 15
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 20

    # Define the table range and add a table with Total Row
    last_column_letter = get_column_letter(ws.max_column)
    table_ref = f"A1:{last_column_letter}{ws.max_row}"
    tab = Table(displayName="ReportTable", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    tab.showTotals = True
    for idx, col in enumerate(tab.tableColumns):
        col.totalsRowLabel = "Total" if idx == 0 else None  # Set only the first column's total label

    ws.add_table(tab)

    wb.save(output)
    output.seek(0)
    return output.getvalue()

# Process data and create calculated columns
def process_data(df):
    df = df.dropna(subset=['Date'])
    
    # Get a complete list of unique dates to ensure all are included in the report
    unique_dates = df['Date'].sort_values().unique()

    # Filters for columns based on keywords in "Description"
    card_to_card_filter = df['Description'].str.contains("انتقال از", na=False)
    fee_filter = df['Description'].str.contains("کارمزد", na=False)
    daily_withdrawal_filter = df['Description'].str.contains("انتقال وجه", na=False)
    snap_deposit_filter = df['Description'].str.contains("مدرن سامانه غذارسان اطلس", na=False)

    # Calculate values for each required column
    card_to_card_sum = df[card_to_card_filter].groupby('Date')['Deposit'].sum().reindex(unique_dates, fill_value=0)
    fee_sum = df[fee_filter].groupby('Date')['Withdrawal'].sum().reindex(unique_dates, fill_value=0)
    daily_withdrawal_sum = df[daily_withdrawal_filter].groupby('Date')['Withdrawal'].sum().reindex(unique_dates, fill_value=0)
    snap_deposit_sum = df[snap_deposit_filter].groupby('Date')['Deposit'].sum().reindex(unique_dates, fill_value=0)
    end_of_day_balance = df.sort_values(['Date', 'Time']).groupby('Date')['Balance'].last().reindex(unique_dates, fill_value=0)

    # Create the report DataFrame
    report = pd.DataFrame({
        'Date': unique_dates,
        'Card_to_Card': card_to_card_sum.values,
        'Fee': fee_sum.values,
        'Daily_Withdrawal': daily_withdrawal_sum.values,
        'Sales': card_to_card_sum.values / 1.1,
        'Tax': card_to_card_sum.values - (card_to_card_sum.values / 1.1),
        'Snap_Deposit': snap_deposit_sum.values,
        'End_of_Day_Balance': end_of_day_balance.values
    })

    # Format values to two decimal places
    report = report[['Date', 'Card_to_Card', 'Sales', 'Tax', 'Fee', 'Daily_Withdrawal', 'End_of_Day_Balance', 'Snap_Deposit']]
    for col in ['Card_to_Card', 'Fee', 'Daily_Withdrawal', 'Sales', 'Tax', 'Snap_Deposit', 'End_of_Day_Balance']:
        report[col] = report[col].apply(lambda x: f"{x:,.2f}")

    # Set final column names in Persian
    report.columns = ['تاریخ', 'کارت به کارت', 'فروش', 'مالیات', 'کارمزد', 'برداشت روز', 'مانده آخر روز', 'واریزی اسنپ']
    
    return report

# Main app setup
def main():
    st.title("Financial Data Report")
    set_custom_style()

    # Display version and update information in the sidebar
    st.sidebar.write(f"**Script Version:** {SCRIPT_VERSION}")
    st.sidebar.write(f"**Last Updated:** {UPDATE_DATE}")

    uploaded_file = st.file_uploader("Choose an Excel or PDF file", type=["xlsx", "pdf"])
    
    if uploaded_file:
        # Encrypt uploaded file content
        encrypted_data = encrypt_file(uploaded_file.getvalue())
        
        # Decrypt file data for processing
        decrypted_data = decrypt_file(encrypted_data)

        if uploaded_file.type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            try:
                df = pd.read_excel(BytesIO(decrypted_data), skiprows=2)
                expected_columns = ['Index', 'Branch Code', 'Branch', 'Date', 'Time', 'Document Number', 
                                    'Receipt Number', 'Check Number', 'Description', 'Withdrawal', 
                                    'Deposit', 'Balance', 'Notes']
                
                if len(df.columns) == len(expected_columns):
                    df.columns = expected_columns
                else:
                    st.error("Uploaded file does not match the expected column structure. Please check the file and try again.")
                    return

            except Exception as e:
                st.error(f"An error occurred while reading the Excel file: {e}")
                return
        elif uploaded_file.type == "application/pdf":
            df = extract_data_from_pdf(BytesIO(decrypted_data))

        report = process_data(df)
        excel_data = create_styled_report(report)

        col1, col2 = st.columns(2)
        with col1:
            if st.button("Preview Report"):
                st.write("### Report Preview")
                st.dataframe(report)

        with col2:
            st.download_button(
                label="Download Report as Excel",
                data=excel_data,
                file_name="Financial_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # Delete all data after processing
        del encrypted_data, decrypted_data, df, report, excel_data

if __name__ == "__main__":
    main()
